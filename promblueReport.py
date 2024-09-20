import os
import pandas as pd
from openpyxl import load_workbook
from xlsxwriter import Workbook
from prometheus_api_client import PrometheusConnect
import configparser
import argparse
import requests
from datetime import datetime, timedelta

def get_latest_config_file(prefix):
    files = [f for f in os.listdir('.') if f.startswith(prefix) and f.endswith('.xlsx')]
    if not files:
        raise FileNotFoundError(f"No files found with prefix '{prefix}'")
    return max(files, key=lambda x: int(x.split('*')[-1].split('.')[0]))

def read_config_file(filename):
    wb = load_workbook(filename)
    ws = wb.active
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[2]  # 세 번째 줄을 칼럼 제목으로 사용
    df = df.iloc[3:]  # 실제 데이터는 네 번째 줄부터 시작
    return df

def load_config(filename):
    config = configparser.ConfigParser()
    config.read(filename)
    return config

def create_workbook_formats(workbook, style_config):
    formats = {}
    for section in style_config.sections():
        if section.startswith('format_'):
            format_dict = dict(style_config[section])
            for key, value in format_dict.items():
                if key in ['bold', 'italic', 'text_wrap']:
                    format_dict[key] = value.lower() == 'true'
                elif key in ['font_size', 'border', 'bottom', 'left', 'right']:
                    format_dict[key] = int(value)
            formats[section] = workbook.add_format(format_dict)
    return formats

def query_prometheus(prom_url, query, start_time, end_time):
    prom = PrometheusConnect(url=prom_url, disable_ssl=True)
    result = prom.custom_query_range(query, start_time=start_time, end_time=end_time, step="1h")
    return result

def get_ollama_feedback(ollama_url, prompt, metrics_dump):
    headers = {'Content-Type': 'application/json'}
    data = {
        "model": "llama2",
        "prompt": f"{prompt}\n\nMetrics dump:\n{metrics_dump}",
        "stream": False
    }
    response = requests.post(ollama_url, headers=headers, json=data)
    if response.status_code == 200:
        return response.json()['response']
    else:
        return "Failed to get feedback from Ollama"

def generate_report(config_df, config, start_time, end_time, output_file):
    workbook = Workbook(output_file)
    worksheet = workbook.add_worksheet()

    # 스타일 설정 로드
    formats = create_workbook_formats(workbook, config)

    # 보고서 제목 작성
    worksheet.write('A1', '서버 점검 보고서', formats['format_title1'])
    worksheet.merge_range('A1:E1', '서버 점검 보고서', formats['format_title1'])

    # 헤더 작성
    headers = ['서버명', 'IP 주소', 'OS', 'CPU 사용률', '메모리 사용률', '디스크 사용률']
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, formats['format_header1'])

    # Prometheus 쿼리 정의
    queries = config['prometheus']['queries'].split(',')
    
    # 데이터 작성
    all_metrics = []
    for row, server in config_df.iterrows():
        worksheet.write(row + 3, 0, server['IT구성정보명'], formats['format_string1'])
        worksheet.write(row + 3, 1, server['IP 주소'], formats['format_string1'])
        worksheet.write(row + 3, 2, server['OS'], formats['format_string1'])

        # Prometheus 쿼리 실행 및 결과 작성
        server_metrics = f"Server: {server['IT구성정보명']} ({server['IP 주소']})\n"
        for col, query in enumerate(queries, start=3):
            result = query_prometheus(config['prometheus']['url'], query.format(ip=server['IP 주소']), start_time, end_time)
            if result:
                value = result[0]['values'][-1][1]  # 마지막 값 사용
                worksheet.write(row + 3, col, float(value), formats['format_stat1'])
                server_metrics += f"{headers[col]}: {value}\n"
            else:
                worksheet.write(row + 3, col, 'N/A', formats['format_string2'])
                server_metrics += f"{headers[col]}: N/A\n"
        all_metrics.append(server_metrics)

    # Ollama 피드백 요청
    metrics_dump = "\n".join(all_metrics)
    ollama_feedback = get_ollama_feedback(config['ollama']['url'], config['ollama']['prompt'], metrics_dump)

    # 종합 의견 작성
    summary_row = len(config_df) + 5
    worksheet.merge_range(f'A{summary_row}:F{summary_row}', '종합 의견', formats['format_title2'])
    worksheet.merge_range(f'A{summary_row+1}:F{summary_row+5}', ollama_feedback, formats['format_string1'])

    workbook.close()

def parse_time_parameter(param):
    now = datetime.now()
    if param == '-1m':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_time = last_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_time = last_month.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif param.endswith('d'):
        days = int(param[:-1])
        start_time = (now - timedelta(days=abs(days))).replace(hour=0, minute=0, second=0, microsecond=0)
        end_time = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif param == '-1':
        start_time = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_time = (now - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif param.endswith('h'):
        hours = int(param[:-1])
        start_time = now - timedelta(hours=abs(hours))
        end_time = now
    else:
        start_time = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_time = now

    return start_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S')

def main():
    parser = argparse.ArgumentParser(description='Generate server inspection report')
    parser.add_argument('--time', type=str, default='', help='Time parameter (e.g., -1m, -30d, -1, -1h)')
    args = parser.parse_args()

    config = load_config('promblueReport.conf')

    config_file = get_latest_config_file(config['files']['config_prefix'])
    config_df = read_config_file(config_file)

    start_time, end_time = parse_time_parameter(args.time)

    generate_report(config_df, config, start_time, end_time, config['files']['output_file'])

if __name__ == "__main__":
    main()